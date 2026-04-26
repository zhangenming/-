from copy import copy

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator

INPUT_PATH = "/Users/zem/zm/未命名文件夹/物流对账(1).xlsx"
OUTPUT_PATH = "/Users/zem/zm/AI/outputs/logistics_reconcile_20260425/物流对账(1)_总表联动版_兼容公式.xlsx"

TOTAL_SHEET = "总表"
SUB_SHEETS = [f"2026.5.{i}" for i in range(1, 32)]
DATA_START_ROW = 8
DATA_END_ROW = 2821
TOTAL_END_ROW = 26996

SHEET_ROW_HELPER_COL = 24   # X
SHEET_COUNT_COL = 25        # Y

TOTAL_HELPER_NAME_START_COL = 24  # X
TOTAL_HELPER_NAME_END_COL = 54    # BB
TOTAL_SEQ_COL = 55                # BC
TOTAL_SHEET_INDEX_COL = 56        # BD
TOTAL_SOURCE_ROW_COL = 57         # BE


def col_letter(index: int) -> str:
    result = []
    while index:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def copy_formula_down(ws, from_row: int, to_row: int, columns: list[int]) -> None:
    for col in columns:
        base = ws.cell(from_row, col)
        for row in range(from_row + 1, to_row + 1):
            target = ws.cell(row, col)
            target.value = Translator(base.value, origin=base.coordinate).translate_formula(target.coordinate)
            target._style = copy(base._style)
            if base.number_format:
                target.number_format = base.number_format
            if base.font:
                target.font = copy(base.font)
            if base.fill:
                target.fill = copy(base.fill)
            if base.border:
                target.border = copy(base.border)
            if base.alignment:
                target.alignment = copy(base.alignment)
            if base.protection:
                target.protection = copy(base.protection)


def write_sub_sheet_helpers(ws) -> None:
    helper_col = col_letter(SHEET_ROW_HELPER_COL)
    count_col = col_letter(SHEET_COUNT_COL)

    ws.cell(1, SHEET_COUNT_COL).value = f'=COUNT(${helper_col}${DATA_START_ROW}:${helper_col}${DATA_END_ROW})'
    ws.cell(DATA_START_ROW, SHEET_ROW_HELPER_COL).value = f'=IF(COUNTA(B{DATA_START_ROW}:G{DATA_START_ROW})=0,"",1)'
    ws.cell(DATA_START_ROW + 1, SHEET_ROW_HELPER_COL).value = (
        f'=IF(COUNTA(B{DATA_START_ROW + 1}:G{DATA_START_ROW + 1})=0,"",MAX(${helper_col}${DATA_START_ROW}:{helper_col}{DATA_START_ROW})+1)'
    )
    copy_formula_down(
        ws,
        from_row=DATA_START_ROW + 1,
        to_row=DATA_END_ROW,
        columns=[SHEET_ROW_HELPER_COL],
    )
    ws.column_dimensions[helper_col].hidden = True
    ws.column_dimensions[count_col].hidden = True


def write_total_sheet_helpers(ws) -> None:
    for row in range(DATA_START_ROW, TOTAL_END_ROW + 1):
        for col in range(1, 8):
            ws.cell(row, col).value = None

    for offset, sheet_name in enumerate(SUB_SHEETS):
        col = TOTAL_HELPER_NAME_START_COL + offset
        ws.cell(1, col).value = sheet_name
        ws.cell(2, col).value = f"='{sheet_name}'!$Y$1"

    first_col_letter = col_letter(TOTAL_HELPER_NAME_START_COL)
    ws.cell(3, TOTAL_HELPER_NAME_START_COL).value = "=0"
    for col in range(TOTAL_HELPER_NAME_START_COL + 1, TOTAL_HELPER_NAME_END_COL + 1):
        prev_col_letter = col_letter(col - 1)
        ws.cell(3, col).value = f"={prev_col_letter}3+{prev_col_letter}2"

    for col in range(TOTAL_HELPER_NAME_START_COL, TOTAL_SOURCE_ROW_COL + 1):
        ws.column_dimensions[col_letter(col)].hidden = True

    ws.cell(DATA_START_ROW, TOTAL_SEQ_COL).value = (
        f'=IF(ROWS($A${DATA_START_ROW}:A{DATA_START_ROW})<=SUM(${first_col_letter}$2:${col_letter(TOTAL_HELPER_NAME_END_COL)}$2),'
        f'ROWS($A${DATA_START_ROW}:A{DATA_START_ROW}),"")'
    )
    ws.cell(DATA_START_ROW, TOTAL_SHEET_INDEX_COL).value = (
        f'=IF($BC{DATA_START_ROW}="","",MATCH($BC{DATA_START_ROW}-1,'
        f'${first_col_letter}$3:${col_letter(TOTAL_HELPER_NAME_END_COL)}$3,1))'
    )
    ws.cell(DATA_START_ROW, TOTAL_SOURCE_ROW_COL).value = (
        f'=IF($BD{DATA_START_ROW}="","",MATCH('
        f'$BC{DATA_START_ROW}-INDEX(${first_col_letter}$3:${col_letter(TOTAL_HELPER_NAME_END_COL)}$3,$BD{DATA_START_ROW}),'
        f'INDIRECT("\'"&INDEX(${first_col_letter}$1:${col_letter(TOTAL_HELPER_NAME_END_COL)}$1,$BD{DATA_START_ROW})&"\'!$X${DATA_START_ROW}:$X${DATA_END_ROW}"),0)'
        f'+{DATA_START_ROW - 1})'
    )

    ws.cell(DATA_START_ROW, 1).value = f'=IF($BC{DATA_START_ROW}="","",$BC{DATA_START_ROW})'
    for col_index, source_col in enumerate(["B", "C", "D", "E", "F", "G"], start=2):
        ws.cell(DATA_START_ROW, col_index).value = (
            f'=IF($BE{DATA_START_ROW}="","",INDIRECT("\'"&INDEX('
            f'${first_col_letter}$1:${col_letter(TOTAL_HELPER_NAME_END_COL)}$1,$BD{DATA_START_ROW})&"\'!{source_col}"&$BE{DATA_START_ROW}))'
        )

    # Keep the existing table appearance on the total sheet while correcting field formats.
    ws.cell(DATA_START_ROW, 1).number_format = "General"
    ws.cell(DATA_START_ROW, 2).number_format = "mm-dd-yy"
    for col in range(3, 8):
        ws.cell(DATA_START_ROW, col).number_format = "General"

    copy_formula_down(
        ws,
        from_row=DATA_START_ROW,
        to_row=TOTAL_END_ROW,
        columns=[1, 2, 3, 4, 5, 6, 7, TOTAL_SEQ_COL, TOTAL_SHEET_INDEX_COL, TOTAL_SOURCE_ROW_COL],
    )


wb = load_workbook(INPUT_PATH)
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True

for sheet_name in SUB_SHEETS:
    write_sub_sheet_helpers(wb[sheet_name])

write_total_sheet_helpers(wb[TOTAL_SHEET])

wb.save(OUTPUT_PATH)
print(OUTPUT_PATH)
